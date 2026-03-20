import io
import os
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import TicketTree, ReportConfig, Comment
from core.gantt import calc_date_range, generate_week_columns, ticket_in_week

HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

STATUS_COLORS = {
    "Backlog": "C9C9C9",
    "To Do": "8DB4E2",
    "In Progress": "00B0F0",
    "In Review": "FFC000",
    "Waiting for Response": "F4B942",
    "Pending Manager Approval": "C39BD3",
    "Blocked": "FF6B6B",
    "On Hold": "F8A487",
    "Done": "00B050",
    "Closed": "70AD47",
    "Cancelled": "E8A598",
    "Duplicate": "E8A598",
}
_FALLBACK_COLOR = "EBEBEB"


def _status_fill(status: str) -> PatternFill:
    color = STATUS_COLORS.get(status, _FALLBACK_COLOR)
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _ticket_id_sort_key(ticket):
    try:
        return int(ticket.id.split("-")[-1])
    except (ValueError, IndexError):
        return 0


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


def _write_header(ws, headers, row=1):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _sheet_tickets(wb, tree: TicketTree):
    ws = wb.create_sheet("Tickets")
    headers = ["ID", "Link", "Name", "Status", "Start Date", "Due Date",
               "Issue Type", "Parent ID", "Parent Name", "Assignee"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    for row_idx, ticket in enumerate(tree.all_tickets(), 2):
        ws.cell(row=row_idx, column=1, value=ticket.id)
        link_cell = ws.cell(row=row_idx, column=2, value=ticket.url)
        link_cell.hyperlink = ticket.url
        link_cell.font = Font(color="0563C1", underline="single")
        ws.cell(row=row_idx, column=3, value=ticket.name)
        ws.cell(row=row_idx, column=4, value=ticket.status)
        ws.cell(row=row_idx, column=5, value=ticket.start_date)
        ws.cell(row=row_idx, column=6, value=ticket.due_date)
        ws.cell(row=row_idx, column=7, value=ticket.issue_type)
        ws.cell(row=row_idx, column=8, value=ticket.parent_id or "")
        ws.cell(row=row_idx, column=9, value=ticket.parent_name or "")
        ws.cell(row=row_idx, column=10, value=ticket.assignee or "")

    _auto_width(ws)


_CLOSED_STATUSES = {"Closed", "Done"}


def _sort_epics_by_due(epics):
    with_date = [e for e in epics if e.due_date is not None]
    without_date = [e for e in epics if e.due_date is None]
    with_date.sort(key=lambda e: e.due_date)
    return with_date + without_date


def _epics_for_gantt(tree: TicketTree, open_statuses):
    """Return closed epics (sorted by due) on top, then open epics (sorted by due)."""
    all_epics = tree.epics()
    closed = _sort_epics_by_due([e for e in all_epics if e.status in _CLOSED_STATUSES])
    open_ = _sort_epics_by_due([e for e in all_epics if e.status in open_statuses])
    return closed + open_


def _write_gantt_headers(ws, headers, week_columns):
    _write_header(ws, headers)
    col_offset = len(headers)
    for i, monday in enumerate(week_columns):
        col = col_offset + i + 1
        friday = monday + timedelta(days=4)
        cell = ws.cell(row=1, column=col, value=friday.strftime("%d/%m"))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _sheet_simplified_gantt(wb, tree: TicketTree, config: ReportConfig):
    ws = wb.create_sheet("Simplified Gantt")
    headers = ["ID", "Name", "Assignee", "Status", "Start Date", "Due Date"]

    sorted_epics = _epics_for_gantt(tree, config.open_statuses)

    min_d, max_d = calc_date_range(sorted_epics)

    if not sorted_epics or min_d is None:
        _write_header(ws, headers)
        ws.freeze_panes = "A2"
        ws.cell(row=2, column=1, value="No epics with dates found")
        _auto_width(ws)
        return

    week_columns = generate_week_columns(min_d, max_d)
    _write_gantt_headers(ws, headers, week_columns)
    col_offset = len(headers)

    for row_idx, epic in enumerate(sorted_epics, 2):
        ws.cell(row=row_idx, column=1, value=epic.id)
        ws.cell(row=row_idx, column=2, value=epic.name)
        ws.cell(row=row_idx, column=3, value=epic.assignee or "")
        status_cell = ws.cell(row=row_idx, column=4, value=epic.status)
        status_cell.fill = _status_fill(epic.status)
        ws.cell(row=row_idx, column=5, value=epic.start_date)
        ws.cell(row=row_idx, column=6, value=epic.due_date)

        fill = _status_fill(epic.status)
        for i, monday in enumerate(week_columns):
            if ticket_in_week(epic, monday):
                ws.cell(row=row_idx, column=col_offset + i + 1).fill = fill

    _auto_width(ws)


def _apply_border_to_row(ws, row_idx: int, total_cols: int, side: Side):
    border = Border(bottom=side)
    for col in range(1, total_cols + 1):
        ws.cell(row=row_idx, column=col).border = border


def _sheet_full_gantt(wb, tree: TicketTree, config: ReportConfig):
    ws = wb.create_sheet("Full Gantt")
    headers = ["ID", "Name", "Assignee", "Status", "Start Date", "Due Date"]

    sorted_epics = _epics_for_gantt(tree, config.open_statuses)

    min_d, max_d = calc_date_range(sorted_epics)

    if not sorted_epics or min_d is None:
        _write_header(ws, headers)
        ws.freeze_panes = "A2"
        ws.cell(row=2, column=1, value="No epics with dates found")
        _auto_width(ws)
        return

    week_columns = generate_week_columns(min_d, max_d)
    _write_gantt_headers(ws, headers, week_columns)
    col_offset = len(headers)
    total_cols = col_offset + len(week_columns)

    thin_side = Side(border_style="thin")
    medium_side = Side(border_style="medium")

    row_idx = 2
    for epic in sorted_epics:
        ws.cell(row=row_idx, column=1, value=epic.id)
        ws.cell(row=row_idx, column=2, value=epic.name)
        ws.cell(row=row_idx, column=3, value=epic.assignee or "")
        status_cell = ws.cell(row=row_idx, column=4, value=epic.status)
        status_cell.fill = _status_fill(epic.status)
        ws.cell(row=row_idx, column=5, value=epic.start_date)
        ws.cell(row=row_idx, column=6, value=epic.due_date)

        epic_fill = _status_fill(epic.status)
        for i, monday in enumerate(week_columns):
            if ticket_in_week(epic, monday):
                ws.cell(row=row_idx, column=col_offset + i + 1).fill = epic_fill

        row_idx += 1

        sorted_children = sorted(epic.children, key=_ticket_id_sort_key)
        for child in sorted_children:
            ws.cell(row=row_idx, column=1, value=f"  {child.id}")
            ws.cell(row=row_idx, column=2, value=child.name)
            ws.cell(row=row_idx, column=3, value=child.assignee or "")
            status_cell = ws.cell(row=row_idx, column=4, value=child.status)
            status_cell.fill = _status_fill(child.status)
            ws.cell(row=row_idx, column=5, value=child.start_date)
            ws.cell(row=row_idx, column=6, value=child.due_date)

            _apply_border_to_row(ws, row_idx, total_cols, thin_side)
            row_idx += 1

            sorted_subtasks = sorted(child.children, key=_ticket_id_sort_key)
            for subtask in sorted_subtasks:
                ws.cell(row=row_idx, column=1, value=f"    {subtask.id}")
                ws.cell(row=row_idx, column=2, value=subtask.name)
                ws.cell(row=row_idx, column=3, value=subtask.assignee or "")
                status_cell = ws.cell(row=row_idx, column=4, value=subtask.status)
                status_cell.fill = _status_fill(subtask.status)
                ws.cell(row=row_idx, column=5, value=subtask.start_date)
                ws.cell(row=row_idx, column=6, value=subtask.due_date)

                _apply_border_to_row(ws, row_idx, total_cols, thin_side)
                row_idx += 1

        # Apply medium border to the last row of the epic group (row_idx - 1)
        _apply_border_to_row(ws, row_idx - 1, total_cols, medium_side)

    _auto_width(ws)


def _filter_comments(comments, root_id, include_root: bool, lookback_days: int):
    cutoff = datetime.now() - timedelta(days=lookback_days)
    filtered = []
    for c in comments:
        if c.created < cutoff:
            continue
        is_root = c.ticket_id == root_id
        if include_root and is_root:
            filtered.append(c)
        elif not include_root and not is_root:
            filtered.append(c)
    filtered.sort(key=lambda c: c.created, reverse=True)
    return filtered


def _sheet_updates(wb, tree: TicketTree, config: ReportConfig):
    ws = wb.create_sheet("Updates")
    headers = ["Ticket ID", "Ticket Name", "Author", "Date", "Comment"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    comments = _filter_comments(tree.comments, tree.root.id, include_root=False, lookback_days=config.lookback_days)

    if not comments:
        ws.cell(row=2, column=1, value="No comments in the last 7 days")
        _auto_width(ws)
        return

    for row_idx, c in enumerate(comments, 2):
        ws.cell(row=row_idx, column=1, value=c.ticket_id)
        ws.cell(row=row_idx, column=2, value=c.ticket_name)
        ws.cell(row=row_idx, column=3, value=c.author)
        ws.cell(row=row_idx, column=4, value=c.created)
        ws.cell(row=row_idx, column=5, value=c.body)

    _auto_width(ws)


def _sheet_weekly_updates(wb, tree: TicketTree, config: ReportConfig):
    ws = wb.create_sheet("Weekly Updates")
    headers = ["Author", "Date", "Comment"]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    comments = _filter_comments(tree.comments, tree.root.id, include_root=True, lookback_days=config.lookback_days)

    if not comments:
        ws.cell(row=2, column=1, value="No comments in the last 7 days")
        _auto_width(ws)
        return

    for row_idx, c in enumerate(comments, 2):
        ws.cell(row=row_idx, column=1, value=c.author)
        ws.cell(row=row_idx, column=2, value=c.created)
        ws.cell(row=row_idx, column=3, value=c.body)

    _auto_width(ws)


def build_excel(tree: TicketTree, config: ReportConfig) -> bytes:
    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    _sheet_tickets(wb, tree)
    _sheet_simplified_gantt(wb, tree, config)
    _sheet_full_gantt(wb, tree, config)
    _sheet_updates(wb, tree, config)
    _sheet_weekly_updates(wb, tree, config)

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    # Save to output directory
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)
    filename = f"{config.report_date.strftime('%Y%m%d')}_{tree.root.id}.xlsx"
    filepath = output_dir / filename
    with open(filepath, "wb") as f:
        f.write(xlsx_bytes)

    return xlsx_bytes
